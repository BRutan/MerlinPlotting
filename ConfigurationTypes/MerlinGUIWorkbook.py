##############################################################################
## MerlinGUIWorkbook.py
##############################################################################
## File gets and handles all of the specified locations 

from DirectoryTypes.FileType import FileType
from DirectoryTypes.FileRow import FileRow
import Exceptions.Fatal as Fatals
from openpyxl import load_workbook
import sys

__all__ = ['MerlinGUIWorkbook']

class MerlinGUIWorkbook(FileType):
    " Class handles all configuration paths stored in the Merlin GUI workbook (Merlin.xlsx) on the Locations sheet. "
    ##########################################################
    ## Constructors:
    ##########################################################
    def __init__(self, isUAT):
        """
        * Overloaded constructor. 
        Inputs:
        * isUAT: True if want to use the UAT path. If not a boolean then use False. 
        """
        (super() if sys.version_info[0] > 2 else super(MerlinGUIWorkbook, self)).__init__(self)
        
        isUAT = (isUAT if isinstance(isUAT) else False)
        self.Name = "Merlin.xlsx"
        # Use hard coded UAT path if necessary:
        self.Path = "//isis/common/Operations/Oberon/Merlin/%s" % ("UAT/Merlin - UAT.xlsm" if isUAT else "Merlin.xlsm")
        
    ##########################################################
    ## Class Methods:
    ##########################################################
    def GetContents(self):
        """
        * Pull in all necessary input and output values from the Merlin gui workbook.
        """
        # Raise exception immediately if file is missing:
        if not self.FileExists(self.Path):
            Fatals.ConfigFilesMissing(callingFunc = 'MerlinGUIWorkbook::GetContents()', missingConfig = self.Name, configPath = self.Path)
        
        # Pull in contents from the workbook:
        try:
            merlinGUIWB = load_workbook(self.Path)
            if 'Locations' not in merlinGUIWB.sheetnames:
                raise Exception(message = 'Locations sheet is missing from Merlin.xlsm.')
            contents = []
            

        except Exception as err:
            raise Fatals.ConfigFilesMissing(callingFunc = 'MerlinGUIWorkbook::GetContents()', missingConfig = self.Name, configPath = self.Path, specific = err.message)
    ##########################################################
    ## Properties:
    ##########################################################
    @property
    def MerlinCurveInput(self):
        " Return path to merlin curves. "
        return self.__MerlinCurveInput
    @property
    def GraphingConfigLocation(self):
        " Return path to the curve graph configuration file (Curve Plotting Configuration.xlsx)."
        return self.__GraphingConfigLocation
    @property 
    def PDFOutput(self):
        " Return the predetermined PDF output location. "
        return self.__PDFOutput
    
    ##########################################################
    ## Setters:
    ##########################################################    
    @GraphingConfigLocation.setter
    def GraphingConfigLocation(self, value):
        """
        * Set the graphing configuration file location (Curve Plotting Configuration.xlsx)/
        """
        if isinstance(value, str):
            if "Curve Plotting Configuration.csv" in value:
                self.__GraphingConfigLocation = value
            else:
                raise ValueError("GraphingConfigLocation must be \<Path>\Curve Plotting configuration.csv")
        else :
            raise ValueError("GraphingConfigLocation must be a string.")
    @PDFOutput.setter
    def PDFOutput(self, value):
        """
        * Set the PDF output path.
        Inputs:
        * value: Expecting a folder string corresponding to where the PDF will be output.
        """
        if isinstance(value, str):
            self.__PDFOutput = FileType.ExtractFolderName(value)
        else:
            raise ValueError("PDFOutput must be a string.")
    @MerlinCurveInput.setter
    def MerlinCurveInput(self, value):
        """
        * Set the merlin curve input path.
        Input:
        * value: Expecting a string filepath with batch time specified. 
        """
        if isinstance(value, str): 
            self.__MerlinCurveInput = ''
            for batch in [ 'AM', 'PM', 'CAPULA' ]:
                if batch in value:
                    self.__MerlinCurveInput = FileType.ExtractFolderName(value)
                    break
            if not self.__MerlinCurveInput:
                raise ValueError('MerlinCurveInput must have batch specified.')
        else:
            raise ValueError('MerlinCurveInput must be a folder path.')
    
    
    

    
        
